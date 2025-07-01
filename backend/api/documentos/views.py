from rest_framework import viewsets
from .models import Documents
from .serializers import DocumentsSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.shortcuts import render
import os
import PyPDF2
import openpyxl
from api.permissions import IsAuthenticatedAndRole


class DocumentsViewSet(viewsets.ModelViewSet):
    queryset = Documents.objects.all()
    serializer_class = DocumentsSerializer
    permission_classes = [IsAuthenticatedAndRole]
    http_method_names = ['get', 'post', 'put', 'delete']
    #required_role = 'admin'


# en desarrollo ( no funcional )
@api_view(['GET'])
def get_document_content(request, pk):
    try:
        document = Documents.objects.get(pk=pk)
        file_path = document.document.path

        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == '.pdf':
            try:
                with open(file_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    content = ''
                    for page in pdf_reader.pages:
                        content += page.extract_text() + '\n'
                html_content = '<p>' + content.replace('\n', '</p><p>') + '</p>'
            except Exception as e:
                return Response({'error': f'Error reading PDF: {str(e)}'}, status=500)

        elif file_extension == '.xlsx' or file_extension == '.xls':
            try:
                workbook = openpyxl.load_workbook(file_path)
                html_content = '<table>'
                for sheet in workbook:
                    html_content += '<tr><th colspan="{}">{}</th></tr>'.format(sheet.max_column, sheet.title)
                    for row in sheet.iter_rows():
                        html_content += '<tr>'
                        for cell in row:
                            html_content += '<td>{}</td>'.format(cell.value if cell.value else '')
                        html_content += '</tr>'
                html_content += '</table>'
            except Exception as e:
                return Response({'error': f'Error reading Excel file: {str(e)}'}, status=500)

        else:
            return Response({'error': 'Unsupported file type.'}, status=400)

        return render(request, 'document_content.html', {'content': html_content})

    except Documents.DoesNotExist:
        return Response({'error': 'Document not found.'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)